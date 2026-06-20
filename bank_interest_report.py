"""
Nepal bank deposit interest-rate announcement extractor.

What it does:
1. Opens Merolagani announcement list.
2. Filters Announcement Type = Interest Rate.
3. Processes only:
   - Commercial Banks
   - Development Bank Limited
4. Keeps latest current-month announcement per bank.
5. Downloads original source document bank-wise.
6. Extracts deposit rates conservatively using OCR/text parsing.
7. Creates formatted Excel report:
   - Interest Rate Summary
   - Notes

Important:
- OCR is conservative. If a value cannot be read confidently, it marks "Unclear"
  or "Not specified" instead of inventing rates.
"""

from __future__ import annotations

import json
import logging
import mimetypes
import os
import re
import sys
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Optional
from urllib.parse import quote, unquote, urljoin, urlparse, urlunparse
from zoneinfo import ZoneInfo

import fitz  # PyMuPDF
import pytesseract
import requests
from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


# ----------------------------
# Configuration
# ----------------------------

BASE_URL = "https://merolagani.com"
LIST_URL = "https://merolagani.com/AnnouncementList.aspx"

TZ = ZoneInfo("Asia/Kathmandu")

TARGET_SECTORS = [
    "Commercial Banks",
    "Development Bank Limited",
]

ANNOUNCEMENT_TYPE = "Interest Rate"

OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", "output"))
DOCS_DIR = OUTPUT_DIR / "latest_bank_files"
REPORT_PATH = OUTPUT_DIR / "interest_rate_report.xlsx"
METADATA_PATH = OUTPUT_DIR / "selected_announcements.json"

OCR_LANG = os.getenv("OCR_LANG", "eng+nep")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/125.0 Safari/537.36"
    )
}

SOURCE_DOC_SELECTORS = [
    "#ctl00_ContentPlaceHolder1_divImage #viewer img",
    "#viewer img",
    ".viewer img",
    ".iviewer_cursor img",
    "#ctl00_ContentPlaceHolder1_divImage img",
    'img[src*="/Uploads/Repository/"]',
    'img[src*="images.merolagani.com"]',
]

RATE_COLUMNS = [
    "Bank",
    "Saving - Min",
    "Saving - Max",
    "Call",
    "Individual FD - Less Than 1 Year Max",
    "Individual FD - 1 Year",
    "Individual FD - More Than 1 Year Max",
    "Institution FD - Less Than 1 Year Max",
    "Institution FD - 1 Year",
    "Institution FD - More Than 1 Year Max",
]

EXCLUDE_KEYWORDS = [
    "loan",
    "lending",
    "base rate",
    "spread",
    "premium",
    "penalty",
    "fee",
    "charge",
    "service charge",
    "overdraft",
    "home loan",
    "auto loan",
    "hire purchase",
    "working capital",
    "margin lending",
    "interest spread",
    "कर्जा",
    "ऋण",
    "शुल्क",
    "जरिवाना",
]


# ----------------------------
# Data classes
# ----------------------------

@dataclass
class Announcement:
    sector: str
    bank_name: str
    title: str
    announcement_date: date
    source_url: str
    ignored_duplicates: list[str] = field(default_factory=list)


@dataclass
class ExtractedRates:
    bank: str
    effective_date: str = "Not specified"
    saving_min: str = "Not specified"
    saving_max: str = "Not specified"
    call: str = "Not specified"
    individual_fd_lt_1y_max: str = "Not specified"
    individual_fd_1y: str = "Not specified"
    individual_fd_gt_1y_max: str = "Not specified"
    institution_fd_lt_1y_max: str = "Not specified"
    institution_fd_1y: str = "Not specified"
    institution_fd_gt_1y_max: str = "Not specified"
    notes: list[str] = field(default_factory=list)
    source_document_url: str = ""
    local_document_path: str = ""


# ----------------------------
# Logging and session
# ----------------------------

def setup_logging() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(OUTPUT_DIR / "run.log", encoding="utf-8"),
        ],
    )


def create_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(HEADERS)

    retries = Retry(
        total=4,
        connect=4,
        read=4,
        backoff_factor=1.2,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "HEAD"],
    )

    adapter = HTTPAdapter(max_retries=retries)
    session.mount("https://", adapter)
    session.mount("http://", adapter)

    return session


# ----------------------------
# Helpers
# ----------------------------

def current_month_bounds() -> tuple[date, date]:
    today = datetime.now(TZ).date()
    start = today.replace(day=1)

    if start.month == 12:
        end = date(start.year + 1, 1, 1)
    else:
        end = date(start.year, start.month + 1, 1)

    return start, end


def safe_filename(value: str, max_len: int = 120) -> str:
    value = re.sub(r"[^\w\s.-]", "", value, flags=re.UNICODE)
    value = re.sub(r"\s+", "_", value.strip())
    return value[:max_len] or "unknown"


def normalize_url(raw_url: str, base_url: str) -> str:
    if not raw_url:
        return ""

    raw_url = raw_url.strip()
    absolute = urljoin(base_url, raw_url)
    parsed = urlparse(absolute)

    path = quote(unquote(parsed.path), safe="/:%")
    query = quote(unquote(parsed.query), safe="=&?/%:+,.-_")

    return urlunparse(
        (
            parsed.scheme,
            parsed.netloc,
            path,
            parsed.params,
            query,
            parsed.fragment,
        )
    )


def parse_announcement_date(text: str) -> Optional[date]:
    text = re.sub(r"\s+", " ", text.strip())

    patterns = [
        r"\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2},\s+\d{4}\b",
        r"\b\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{4}\b",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if not match:
            continue

        raw = match.group(0).replace(".", "")

        for fmt in ("%b %d, %Y", "%B %d, %Y", "%d %b %Y", "%d %B %Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                pass

    return None


def extract_bank_name_from_title(title: str) -> str:
    title = re.sub(r"\s+", " ", title).strip()

    patterns = [
        r"^(.*?)\s+has\s+(?:published|made|issued|notified|urged|announced)",
        r"^(.*?)\s+is\s+(?:going|scheduled)",
        r"^(.*?)\s+made\s+correction",
        r"-\s*([^()-]*(?:Bank|Bikas Bank|Development Bank)[^()]*)",
    ]

    for pattern in patterns:
        match = re.search(pattern, title, flags=re.I)

        if match:
            name = match.group(1).strip(" -")
            name = re.sub(r"\s*\(Former.*?\)", "", name, flags=re.I).strip()

            if len(name) >= 4:
                return name

    fallback = re.sub(r"\s+has\s+.*$", "", title, flags=re.I).strip()
    return fallback[:120] if fallback else "Unknown Bank"


def is_current_month(d: date) -> bool:
    start, end = current_month_bounds()
    return start <= d < end


def is_probably_interest_rate(title: str) -> bool:
    text = title.lower()
    return "interest rate" in text or "new interest" in text or "deposit" in text


# ----------------------------
# Announcement collection
# ----------------------------

def select_dropdown_by_option_text(page, option_text: str) -> bool:
    selects = page.locator("select")
    count = selects.count()

    for i in range(count):
        select = selects.nth(i)

        try:
            options = select.locator("option").all_inner_texts()
        except Exception:
            continue

        cleaned = [re.sub(r"\s+", " ", option).strip() for option in options]

        if option_text in cleaned:
            select.select_option(label=option_text)
            return True

    return False


def click_search(page) -> None:
    """
    Click Merolagani Announcement Search button.

    Real Search button HTML:
    <a id="ctl00_ContentPlaceHolder1_lbtnSearch" class="btn btn-primary">Search</a>

    This avoids the old issue where Playwright selected a hidden:
    a:has-text('Search')
    """

    search_selector = "#ctl00_ContentPlaceHolder1_lbtnSearch"

    page.wait_for_load_state("domcontentloaded")
    page.set_viewport_size({"width": 1600, "height": 1200})
    page.wait_for_timeout(1000)

    page.wait_for_selector(search_selector, state="attached", timeout=30000)

    search_button = page.locator(search_selector).first

    try:
        if search_button.is_visible():
            search_button.scroll_into_view_if_needed(timeout=5000)
            search_button.click(timeout=15000)
        else:
            raise RuntimeError("Search button exists but is not visible.")

    except Exception as exc:
        logging.warning("Normal Search click failed. Using ASP.NET postback fallback: %s", exc)

        page.evaluate(
            """
            () => {
                const el = document.querySelector("#ctl00_ContentPlaceHolder1_lbtnSearch");

                if (!el) {
                    throw new Error("Search button not found.");
                }

                try {
                    if (typeof showProcessing === "function") {
                        showProcessing();
                    }
                } catch (e) {}

                try {
                    if (typeof webEngageTrackEvent === "function") {
                        webEngageTrackEvent("Announcement Search");
                    }
                } catch (e) {}

                if (typeof __doPostBack === "function") {
                    __doPostBack("ctl00$ContentPlaceHolder1$lbtnSearch", "");
                } else {
                    el.click();
                }
            }
            """
        )

    page.wait_for_timeout(2000)

    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except PlaywrightTimeoutError:
        pass


def click_load_more(page) -> bool:
    selectors = [
        "text=Load More",
        "button:has-text('Load More')",
        "a:has-text('Load More')",
    ]

    for selector in selectors:
        locator = page.locator(selector)

        if locator.count() > 0 and locator.first.is_visible():
            before = page.content()
            locator.first.click()
            page.wait_for_timeout(2000)
            after = page.content()

            return before != after

    return False


def parse_announcements_from_html(html: str, sector: str) -> list[Announcement]:
    soup = BeautifulSoup(html, "html.parser")
    results: list[Announcement] = []

    links = []

    for a in soup.find_all("a", href=True):
        title = a.get_text(" ", strip=True)
        href = a.get("href", "")

        if len(title) < 10:
            continue

        href_low = href.lower()
        title_low = title.lower()

        if (
            "announcement" not in href_low
            and "announcementdetail" not in href_low
            and "published" not in title_low
            and "interest rate" not in title_low
        ):
            continue

        if not is_probably_interest_rate(title):
            continue

        links.append(a)

    for a in links:
        title = re.sub(r"\s+", " ", a.get_text(" ", strip=True)).strip()
        source_url = normalize_url(a["href"], LIST_URL)

        date_found = None

        node: Optional[Tag] = a

        for _ in range(5):
            if node is None:
                break

            text = node.get_text(" ", strip=True)
            date_found = parse_announcement_date(text)

            if date_found:
                break

            node = node.parent if isinstance(node.parent, Tag) else None

        if not date_found:
            prev_texts = []
            prev = a.previous_element

            for _ in range(40):
                if prev is None:
                    break

                if isinstance(prev, str):
                    prev_texts.append(prev)
                elif isinstance(prev, Tag):
                    prev_texts.append(prev.get_text(" ", strip=True))

                date_found = parse_announcement_date(" ".join(reversed(prev_texts)))

                if date_found:
                    break

                prev = prev.previous_element

        if not date_found:
            logging.warning("Skipping announcement with missing date: %s", title)
            continue

        if not is_current_month(date_found):
            continue

        bank = extract_bank_name_from_title(title)

        results.append(
            Announcement(
                sector=sector,
                bank_name=bank,
                title=title,
                announcement_date=date_found,
                source_url=source_url,
            )
        )

    return results


def collect_current_month_announcements() -> list[Announcement]:
    logging.info("Collecting current-month Interest Rate announcements from Merolagani.")

    selected: dict[tuple[str, str], Announcement] = {}

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(user_agent=HEADERS["User-Agent"])

        try:
            for sector in TARGET_SECTORS:
                logging.info("Processing sector: %s", sector)

                page.goto(LIST_URL, wait_until="domcontentloaded", timeout=60000)

                try:
                    page.wait_for_load_state("networkidle", timeout=15000)
                except PlaywrightTimeoutError:
                    pass

                if not select_dropdown_by_option_text(page, sector):
                    raise RuntimeError(f"Sector option not found: {sector}")

                if not select_dropdown_by_option_text(page, ANNOUNCEMENT_TYPE):
                    raise RuntimeError(f"Announcement Type option not found: {ANNOUNCEMENT_TYPE}")

                click_search(page)

                sector_records: list[Announcement] = []

                for _ in range(12):
                    html = page.content()
                    records = parse_announcements_from_html(html, sector)
                    sector_records = records

                    if records:
                        oldest = min(record.announcement_date for record in records)
                        month_start, _ = current_month_bounds()

                        if oldest < month_start:
                            break

                    if not click_load_more(page):
                        break

                for ann in sector_records:
                    key = (sector, ann.bank_name.lower())
                    existing = selected.get(key)

                    if existing is None:
                        selected[key] = ann
                    else:
                        if ann.announcement_date > existing.announcement_date:
                            ann.ignored_duplicates.append(
                                f"Ignored older duplicate: {existing.announcement_date.isoformat()} | {existing.title}"
                            )
                            selected[key] = ann
                        else:
                            existing.ignored_duplicates.append(
                                f"Ignored older duplicate: {ann.announcement_date.isoformat()} | {ann.title}"
                            )

        finally:
            browser.close()

    final = sorted(
        selected.values(),
        key=lambda item: (item.sector, item.bank_name.lower(), item.announcement_date),
    )

    if not final:
        raise RuntimeError(
            "No current-month Interest Rate announcements found for Commercial Banks "
            "or Development Bank Limited. Check Merolagani availability or page structure."
        )

    METADATA_PATH.parent.mkdir(parents=True, exist_ok=True)
    METADATA_PATH.write_text(
        json.dumps([asdict(item) for item in final], ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )

    logging.info("Selected %s latest bank-wise announcements.", len(final))
    return final


# ----------------------------
# Source document extraction
# ----------------------------

def find_source_document_url(html: str, detail_url: str) -> str:
    soup = BeautifulSoup(html, "html.parser")

    for selector in SOURCE_DOC_SELECTORS:
        for img in soup.select(selector):
            for attr in ["src", "data-src", "data-original", "data-zoom-image"]:
                raw = img.get(attr)

                if raw:
                    return normalize_url(raw, detail_url)

    for a in soup.find_all("a", href=True):
        href = a["href"]

        if re.search(r"\.(pdf|jpg|jpeg|png|webp)(?:\?|$)", href, flags=re.I):
            return normalize_url(href, detail_url)

        if "/Uploads/Repository/" in href:
            return normalize_url(href, detail_url)

    for img in soup.find_all("img"):
        raw = img.get("src") or ""

        if "/Uploads/Repository/" in raw or "images.merolagani.com" in raw:
            return normalize_url(raw, detail_url)

    return ""


def fetch_detail_and_document_url(session: requests.Session, ann: Announcement) -> str:
    logging.info("Opening detail page: %s", ann.source_url)

    response = session.get(ann.source_url, timeout=45)
    response.raise_for_status()

    doc_url = find_source_document_url(response.text, ann.source_url)

    if not doc_url:
        logging.warning("No source document found for %s", ann.bank_name)
        return ""

    return doc_url


def get_extension_from_response(url: str, response: requests.Response) -> str:
    content_type = response.headers.get("Content-Type", "").split(";")[0].strip().lower()
    ext = mimetypes.guess_extension(content_type) or ""

    if ext in [".jpe"]:
        ext = ".jpg"

    parsed_ext = Path(urlparse(url).path).suffix.lower()

    if parsed_ext in [".pdf", ".jpg", ".jpeg", ".png", ".webp"]:
        return parsed_ext

    return ext or ".bin"


def download_source_document(
    session: requests.Session,
    ann: Announcement,
    source_document_url: str,
) -> str:
    sector_dir = DOCS_DIR / safe_filename(ann.sector)
    bank_dir = sector_dir / safe_filename(ann.bank_name)
    bank_dir.mkdir(parents=True, exist_ok=True)

    if not source_document_url:
        return ""

    response = session.get(source_document_url, timeout=90)
    response.raise_for_status()

    ext = get_extension_from_response(source_document_url, response)
    filename = f"{ann.announcement_date.isoformat()}_{safe_filename(ann.bank_name)}{ext}"
    path = bank_dir / filename

    path.write_bytes(response.content)

    logging.info("Saved source document: %s", path)

    return str(path)


# ----------------------------
# Text/OCR extraction
# ----------------------------

def extract_text_from_pdf(path: Path) -> str:
    text_parts: list[str] = []

    doc = fitz.open(path)

    try:
        for page in doc:
            txt = page.get_text("text") or ""

            if txt.strip():
                text_parts.append(txt)
    finally:
        doc.close()

    text = "\n".join(text_parts).strip()

    if len(text) >= 80:
        return text

    logging.info("PDF has little embedded text. Running OCR: %s", path)

    ocr_parts: list[str] = []

    doc = fitz.open(path)

    try:
        for page_index, page in enumerate(doc):
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            tmp_img = path.with_suffix(f".page_{page_index + 1}.png")
            pix.save(str(tmp_img))

            try:
                img = Image.open(tmp_img)
                ocr_parts.append(pytesseract.image_to_string(img, lang=OCR_LANG))
            finally:
                tmp_img.unlink(missing_ok=True)
    finally:
        doc.close()

    return "\n".join(ocr_parts)


def extract_text_from_image(path: Path) -> str:
    img = Image.open(path)
    return pytesseract.image_to_string(img, lang=OCR_LANG)


def extract_document_text(local_path: str) -> str:
    if not local_path:
        return ""

    path = Path(local_path)
    suffix = path.suffix.lower()

    if suffix == ".pdf":
        return extract_text_from_pdf(path)

    if suffix in [".jpg", ".jpeg", ".png", ".webp"]:
        return extract_text_from_image(path)

    return ""


# ----------------------------
# Rate parsing
# ----------------------------

def normalize_text(text: str) -> str:
    text = text.replace("％", "%")
    text = text.replace("–", "-").replace("—", "-")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)

    return text.strip()


def has_excluded_context(line: str) -> bool:
    low = line.lower()
    return any(keyword in low for keyword in EXCLUDE_KEYWORDS)


def extract_rates_from_line(line: str) -> list[float]:
    found: list[float] = []

    for match in re.finditer(r"(?<!\d)(\d{1,2}(?:\.\d{1,2})?)\s*%?", line):
        try:
            value = float(match.group(1))
        except ValueError:
            continue

        if 0 <= value <= 25:
            found.append(value)

    return found


def rate_text(value: Optional[float]) -> str:
    if value is None:
        return "Not specified"

    text = f"{value:.2f}".rstrip("0").rstrip(".")

    return f"{text}%"


def max_rate_text(values: list[float]) -> str:
    return rate_text(max(values)) if values else "Not specified"


def exact_or_max_rate_text(values: list[float]) -> str:
    return rate_text(max(values)) if values else "Not specified"


def detect_effective_date(text: str) -> str:
    patterns = [
        r"(?:effective from|with effect from|applicable from)\s*[:\-]?\s*([A-Za-z]+\s+\d{1,2},?\s+\d{4})",
        r"(?:effective from|with effect from|applicable from)\s*[:\-]?\s*(\d{1,2}\s+[A-Za-z]+\s+\d{4})",
        r"(?:effective from|with effect from|applicable from)\s*[:\-]?\s*([A-Za-z]+\s+\d{1,2})",
        r"(?:effective from|with effect from|applicable from)\s*[:\-]?\s*([A-Za-z]+\s+\d{1,2},?\s+\d{4}|Ashad\s+\d{1,2},?\s+\d{4})",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)

        if match:
            return match.group(1).strip()

    return "Not specified"


def detect_bank_name_from_document(text: str, fallback: str) -> str:
    lines = [re.sub(r"\s+", " ", item).strip() for item in text.splitlines()]
    lines = [item for item in lines if item]

    for line in lines[:40]:
        match = re.search(
            r"([A-Z][A-Za-z&.\s]+(?:Bank|Bikas Bank|Development Bank)\s+Limited)",
            line,
            flags=re.I,
        )

        if match:
            return re.sub(r"\s+", " ", match.group(1)).strip()

    return fallback


def line_has_any(line: str, keywords: list[str]) -> bool:
    low = line.lower()
    return any(keyword.lower() in low for keyword in keywords)


def bucket_for_fd_line(line: str) -> Optional[str]:
    low = line.lower()

    more_patterns = [
        "above 1 year",
        "above one year",
        "more than 1 year",
        "more than one year",
        "over 1 year",
        ">1 year",
        "1 year and above",
        "one year and above",
        "beyond 1 year",
    ]

    less_patterns = [
        "less than 1 year",
        "less than one year",
        "below 1 year",
        "below one year",
        "up to 1 year",
        "upto 1 year",
        "up-to 1 year",
        "<1 year",
        "3 month",
        "6 month",
        "9 month",
        "91 days",
        "180 days",
        "181 days",
        "270 days",
    ]

    exact_patterns = [
        r"\b1\s*year\b",
        r"\bone\s*year\b",
        r"\b12\s*months?\b",
    ]

    if any(pattern in low for pattern in more_patterns):
        return "gt_1y"

    if any(pattern in low for pattern in less_patterns):
        return "lt_1y"

    if any(re.search(pattern, low) for pattern in exact_patterns):
        if not any(
            marker in low
            for marker in [
                "above",
                "more",
                "over",
                "below",
                "less",
                "up to",
                "upto",
                "and above",
            ]
        ):
            return "one_y"

    return None


def parse_deposit_rates(text: str, fallback_bank: str) -> ExtractedRates:
    text = normalize_text(text)

    rates = ExtractedRates(bank=detect_bank_name_from_document(text, fallback_bank))
    rates.effective_date = detect_effective_date(text)

    if not text.strip():
        rates.notes.append("Source document could not be read; all rates marked Not specified.")
        return rates

    raw_lines = [re.sub(r"\s+", " ", item).strip() for item in text.splitlines()]
    lines = [item for item in raw_lines if item]

    saving_values: list[float] = []
    call_values: list[float] = []

    individual = {
        "lt_1y": [],
        "one_y": [],
        "gt_1y": [],
    }

    institution = {
        "lt_1y": [],
        "one_y": [],
        "gt_1y": [],
    }

    current_context: Optional[str] = None

    for line in lines:
        low = line.lower()

        if has_excluded_context(line):
            continue

        if line_has_any(line, ["saving", "savings", "बचत"]):
            current_context = "saving"
        elif line_has_any(line, ["call deposit", "call account", "call", "कल"]):
            current_context = "call"
        elif line_has_any(line, ["institutional fixed", "institution fd", "institution", "संस्थागत"]):
            current_context = "institution_fd"
        elif line_has_any(line, ["individual fixed", "individual fd", "personal fixed", "individual", "व्यक्तिगत"]):
            current_context = "individual_fd"
        elif line_has_any(line, ["fixed deposit", "term deposit", "fd", "मुद्दती"]):
            current_context = current_context or "fixed_deposit"

        values = extract_rates_from_line(line)

        if not values:
            continue

        if current_context == "saving" or line_has_any(line, ["saving", "savings", "बचत"]):
            saving_values.extend(values)
            continue

        if current_context == "call" or line_has_any(line, ["call deposit", "call account", "call", "कल"]):
            call_values.extend(values)
            continue

        bucket = bucket_for_fd_line(line)

        is_institution = line_has_any(
            line,
            ["institution", "institutional", "organization", "corporate", "संस्थागत"],
        )

        is_individual = line_has_any(
            line,
            ["individual", "personal", "natural person", "व्यक्तिगत"],
        )

        if bucket and is_individual and is_institution and len(values) >= 2:
            ind_pos = min(
                [low.find(item) for item in ["individual", "personal"] if low.find(item) >= 0]
                or [9999]
            )

            inst_pos = min(
                [
                    low.find(item)
                    for item in ["institution", "institutional", "organization", "corporate"]
                    if low.find(item) >= 0
                ]
                or [9999]
            )

            if ind_pos <= inst_pos:
                individual[bucket].append(values[0])
                institution[bucket].append(values[1])
            else:
                institution[bucket].append(values[0])
                individual[bucket].append(values[1])

            continue

        if bucket:
            if is_institution or current_context == "institution_fd":
                institution[bucket].extend(values)
            elif is_individual or current_context == "individual_fd":
                individual[bucket].extend(values)
            elif current_context == "fixed_deposit":
                rates.notes.append(
                    f"FD line found but individual/institution category unclear: {line[:180]}"
                )

            continue

    unique_saving = sorted(set(saving_values))

    if unique_saving:
        rates.saving_min = rate_text(unique_saving[0])

        if len(unique_saving) == 1:
            rates.saving_max = rate_text(unique_saving[0])
            rates.notes.append("Only one saving rate found; used same value for Saving Min and Saving Max.")
        else:
            rates.saving_max = rate_text(unique_saving[-2])
            rates.notes.append(
                "Saving Max rule applied: excluded the single highest saving rate and used the next highest rate."
            )
    else:
        if re.search(r"saving|savings|बचत", text, flags=re.I):
            rates.saving_min = "Unclear"
            rates.saving_max = "Unclear"
            rates.notes.append("Saving section detected but rates were not readable confidently.")
        else:
            rates.notes.append("Saving deposit rate not specified/readable in extracted text.")

    if call_values:
        rates.call = max_rate_text(call_values)
    else:
        if re.search(r"call|कल", text, flags=re.I):
            rates.call = "Unclear"
            rates.notes.append("Call deposit section detected but rate was not readable confidently.")
        else:
            rates.notes.append("Call deposit rate not specified/readable in extracted text.")

    rates.individual_fd_lt_1y_max = max_rate_text(individual["lt_1y"])
    rates.individual_fd_1y = exact_or_max_rate_text(individual["one_y"])
    rates.individual_fd_gt_1y_max = max_rate_text(individual["gt_1y"])

    rates.institution_fd_lt_1y_max = max_rate_text(institution["lt_1y"])
    rates.institution_fd_1y = exact_or_max_rate_text(institution["one_y"])
    rates.institution_fd_gt_1y_max = max_rate_text(institution["gt_1y"])

    fd_fields = [
        ("Individual FD <1 Year", rates.individual_fd_lt_1y_max),
        ("Individual FD 1 Year", rates.individual_fd_1y),
        ("Individual FD >1 Year", rates.individual_fd_gt_1y_max),
        ("Institution FD <1 Year", rates.institution_fd_lt_1y_max),
        ("Institution FD 1 Year", rates.institution_fd_1y),
        ("Institution FD >1 Year", rates.institution_fd_gt_1y_max),
    ]

    for label, value in fd_fields:
        if value == "Not specified":
            rates.notes.append(f"{label} not specified/readable in extracted text.")

    return rates


# ----------------------------
# Excel report
# ----------------------------

def apply_border(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.border = border


def auto_adjust_widths(ws, max_width: int = 60) -> None:
    for column_cells in ws.columns:
        length = 0
        col_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))

        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), max_width)


def build_excel_report(records: list[tuple[Announcement, ExtractedRates]]) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "Interest Rate Summary"

    notes = wb.create_sheet("Notes")

    ws["A1"] = "Bank"
    ws["A2"] = "Bank"

    ws.merge_cells("B1:C1")
    ws["B1"] = "Saving"
    ws["B2"] = "Min"
    ws["C2"] = "Max"

    ws["D1"] = "Call"
    ws["D2"] = "Call"

    ws.merge_cells("E1:G1")
    ws["E1"] = "Individual FD"
    ws["E2"] = "<1 Year Max"
    ws["F2"] = "1 Year"
    ws["G2"] = ">1 Year Max"

    ws.merge_cells("H1:J1")
    ws["H1"] = "Institution FD"
    ws["H2"] = "<1 Year Max"
    ws["I2"] = "1 Year"
    ws["J2"] = ">1 Year Max"

    data_start_row = 3

    for row_idx, (ann, extracted_rates) in enumerate(records, start=data_start_row):
        values = [
            extracted_rates.bank,
            extracted_rates.saving_min,
            extracted_rates.saving_max,
            extracted_rates.call,
            extracted_rates.individual_fd_lt_1y_max,
            extracted_rates.individual_fd_1y,
            extracted_rates.individual_fd_gt_1y_max,
            extracted_rates.institution_fd_lt_1y_max,
            extracted_rates.institution_fd_1y,
            extracted_rates.institution_fd_gt_1y_max,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = str(value)
            cell.number_format = "@"

    notes_headers = [
        "Bank",
        "Effective Date",
        "Source Announcement Date",
        "Source URL",
        "Source Document URL",
        "Notes / Assumptions",
    ]

    notes.append(notes_headers)

    for ann, extracted_rates in records:
        note_parts = list(extracted_rates.notes)

        if ann.ignored_duplicates:
            note_parts.extend(ann.ignored_duplicates)

        if not extracted_rates.source_document_url:
            note_parts.append("Source document URL could not be accessed/found.")

        notes.append(
            [
                extracted_rates.bank,
                extracted_rates.effective_date,
                ann.announcement_date.isoformat(),
                ann.source_url,
                extracted_rates.source_document_url or "Not specified",
                " | ".join(note_parts) if note_parts else "No material assumptions noted.",
            ]
        )

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    subheader_fill = PatternFill("solid", fgColor="EAF4FB")
    header_font = Font(bold=True)

    for sheet in [ws, notes]:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for cell in ws[2]:
        cell.font = header_font
        cell.fill = subheader_fill

    for cell in notes[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:J{max(ws.max_row, 2)}"

    notes.freeze_panes = "A2"
    notes.auto_filter.ref = f"A1:F{max(notes.max_row, 1)}"

    apply_border(ws, 1, ws.max_row, 1, 10)
    apply_border(notes, 1, notes.max_row, 1, 6)

    auto_adjust_widths(ws)
    auto_adjust_widths(notes)

    notes.column_dimensions["D"].width = 55
    notes.column_dimensions["E"].width = 55
    notes.column_dimensions["F"].width = 80

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb.save(REPORT_PATH)

    logging.info("Excel report saved: %s", REPORT_PATH)


# ----------------------------
# Verification and final output
# ----------------------------

def missing_or_unclear_banks(records: list[tuple[Announcement, ExtractedRates]]) -> list[str]:
    flagged = []

    for ann, extracted_rates in records:
        values = [
            extracted_rates.saving_min,
            extracted_rates.saving_max,
            extracted_rates.call,
            extracted_rates.individual_fd_lt_1y_max,
            extracted_rates.individual_fd_1y,
            extracted_rates.individual_fd_gt_1y_max,
            extracted_rates.institution_fd_lt_1y_max,
            extracted_rates.institution_fd_1y,
            extracted_rates.institution_fd_gt_1y_max,
        ]

        if any(value in ["Not specified", "Unclear"] for value in values):
            flagged.append(extracted_rates.bank)

    return sorted(set(flagged))


def verify(records: list[tuple[Announcement, ExtractedRates]]) -> None:
    bank_keys = [(ann.sector, extracted_rates.bank.lower()) for ann, extracted_rates in records]

    if len(bank_keys) != len(set(bank_keys)):
        raise RuntimeError("Verification failed: duplicate bank found in summary records.")

    for ann, _ in records:
        if ann.sector not in TARGET_SECTORS:
            raise RuntimeError(f"Verification failed: invalid sector included: {ann.sector}")

        if not is_current_month(ann.announcement_date):
            raise RuntimeError(f"Verification failed: non-current-month announcement included: {ann.title}")

    if not REPORT_PATH.exists():
        raise RuntimeError("Verification failed: Excel report was not created.")


def print_final_response(records: list[tuple[Announcement, ExtractedRates]]) -> None:
    commercial_count = sum(1 for ann, _ in records if ann.sector == "Commercial Banks")
    development_count = sum(1 for ann, _ in records if ann.sector == "Development Bank Limited")
    flagged = missing_or_unclear_banks(records)

    print()
    print("✅ Report created successfully")
    print(
        f"📊 Number of banks compiled: Commercial Banks = {commercial_count}; "
        f"Development Banks = {development_count}; Total = {len(records)}"
    )

    if flagged:
        print("⚠️ Banks with missing or unclear values:")

        for bank in flagged:
            print(f" - {bank}")
    else:
        print("⚠️ Banks with missing or unclear values: None")

    print(f"📎 Excel file: {REPORT_PATH}")
    print(f"📁 Bank-wise source files: {DOCS_DIR}")


# ----------------------------
# Main
# ----------------------------

def main() -> int:
    setup_logging()

    session = create_session()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    DOCS_DIR.mkdir(parents=True, exist_ok=True)

    try:
        announcements = collect_current_month_announcements()
    except Exception as exc:
        logging.exception("Announcement list unavailable or collection failed.")
        print(f"ERROR: Announcement list page unavailable or could not be processed: {exc}")
        return 1

    records: list[tuple[Announcement, ExtractedRates]] = []

    for ann in announcements:
        logging.info("Extracting rates for: %s | %s", ann.sector, ann.bank_name)

        extracted = ExtractedRates(bank=ann.bank_name)

        try:
            doc_url = fetch_detail_and_document_url(session, ann)
            extracted.source_document_url = doc_url

            if not doc_url:
                extracted.notes.append("Source document could not be found on announcement detail page.")
            else:
                local_path = download_source_document(session, ann, doc_url)
                extracted.local_document_path = local_path

                doc_text = extract_document_text(local_path)
                parsed = parse_deposit_rates(doc_text, ann.bank_name)

                parsed.source_document_url = doc_url
                parsed.local_document_path = local_path

                extracted = parsed

        except Exception as exc:
            logging.exception("Failed to process bank: %s", ann.bank_name)
            extracted.notes.append(f"Source document issue encountered: {exc}")
            extracted.source_document_url = extracted.source_document_url or "Not specified"

        records.append((ann, extracted))

    build_excel_report(records)
    verify(records)
    print_final_response(records)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
